from openpyxl import load_workbook
from openpyxl import Workbook
from common.info import InfoPart


class ManageExcel(object):
    '''excel处理'''

    def __init__(self):
        self.wb = load_workbook(InfoPart().test_data_path)
        self.sheet_name = InfoPart().sheet_name
        self.mode = InfoPart().mode
        self.run_list = InfoPart().run_list

    def data_treat(self):
        # 数据处理，mode为all，则全部跑，part则依据run_list来跑
        all_data = {}
        for sheet_name in self.sheet_name:
            # print(sheet_name)
            all_data[sheet_name] = self.read_data(
                sheet_name)  # 将excel中的所有数据输出出来
        if self.mode == 'all_run':  # 判定mode是all_run,则全部跑
            test_data = all_data

        else:
            # pass
            test_data = {}
            for key, values in self.run_list.items(
            ):  # 读取配置文件中runlist的key和value，即表名，caseid
                data = []
                if values == 'all':  # 若为all则都跑，从alldata将该sheet那么下的所有case都取出来
                    # print(key)
                    # data.append(alldata[key][0])
                    test_data[key] = all_data[key]
                else:  # 配置信息不是all，而是part的情况下，取对应的case跑
                    for value in values:  # 遍历caseid
                        # 取出对应的caseid，由于list中是从0开始取的，故value需要减1
                        data.append(all_data[key][value - 1])
                    test_data[key] = data
        return test_data

    def read_data(self, sheet_name):
        # 获取最大列
        max_row = self.wb[sheet_name].max_row
        # 获取最大列
        max_column = self.wb[sheet_name].max_column
        test_data = []
        for row in range(2, max_row + 1):  # 遍历行
            dict1 = {}
            for column in range(1, max_column + 1):  # 遍历列
                # 表的第一行值作为title(key)，存储第二行开始的所有数据
                dict1[self.wb[sheet_name].cell(1, column).value] = self.wb[sheet_name].cell(
                    row, column).value
            test_data.append(dict1)
        return test_data

    def write_back(self, sheet_name, row, test_result, actual_result):
        # 将测试结果(Pass/Fail),实际结果(接口的响应体),写回到case文件中去
        self.wb[sheet_name].cell(
            row,
            self.wb[sheet_name].max_column -
            1).value = test_result
        self.wb[sheet_name].cell(
            row,
            self.wb[sheet_name].max_column -
            2).value = actual_result
        self.save_excel()

    def save_excel(self):
        self.wb.save(InfoPart().test_data_path)


if __name__ == "__main__":
    me = ManageExcel()
    # print(me.data_treat())
    print(me.read_data('ddts'))
    # print(me.read_data(me.sheet_name[0]))
